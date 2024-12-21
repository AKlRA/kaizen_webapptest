from datetime import datetime, timedelta
import json
import os
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from .forms import RegisterForm, KaizenSheetForm, HandwrittenKaizenForm
from .models import HorizontalDeployment, KaizenSheet, Profile, Image
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse, HttpResponseForbidden
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage  # Rename to avoid confusion
from openpyxl.utils import range_boundaries, get_column_letter
from django.utils import timezone
from django.db import models, transaction
from django.views.decorators.csrf import ensure_csrf_cookie
from .models import KaizenCoordinator
from django.views.decorators.http import require_POST



# Home view
def index(request):
    if User.objects.exists():
        return redirect('login')
    else:
        return redirect('register')

def register(request):
    if request.method == 'POST':
        # Add debug prints
        print("POST data:", request.POST)
        
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        password = request.POST.get('password')  
        role = request.POST.get('role')
        department = request.POST.get('department')
        
        # Print received values
        print(f"""
        user_id: {user_id}
        username: {username} 
        password: {password}
        role: {role}
        department: {department}
        """)

        # Check field names match form
        if not all([user_id, username, password, role]):
            print("Missing required fields:")
            if not user_id: print("- user_id")
            if not username: print("- username") 
            if not password: print("- password")
            if not role: print("- role")
            return JsonResponse({'success': False, 'error': 'All fields are required'})

        # Check if user_id already exists
        if User.objects.filter(profile__employee_id=user_id).exists():
            return JsonResponse({'success': False, 'error': 'Employee ID already exists'})

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': 'Username already exists'})

        # Check coordinator/finance limit
        if role == 'coordinator':
            if User.objects.filter(profile__user_type='coordinator').count() >= 2:
                return JsonResponse({'success': False, 'error': 'Maximum coordinator limit reached'})
        elif role == 'finance':
            if User.objects.filter(profile__user_type='finance').count() >= 2:
                return JsonResponse({'success': False, 'error': 'Maximum finance users limit reached'})

        # Validate department for employee and hod
        if role in ['employee', 'hod'] and not department:
            return JsonResponse({'success': False, 'error': 'Department is required for employees and HODs'})

        try:
            with transaction.atomic():
                user = User.objects.create_user(username=username, password=password)
                Profile.objects.create(
                    user=user,
                    user_type=role,
                    department=department if role in ['employee', 'hod'] else None,
                    employee_id=user_id
                )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    # Get all departments from DEPARTMENT_CHOICES
    departments = [dept[0] for dept in Profile.DEPARTMENT_CHOICES]
    return render(request, 'dashboard/register.html', {'departments': departments})
# Login view
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            profile = Profile.objects.get(user=user)
            if profile.user_type == 'employee':
                return redirect('employee_dashboard')
            elif profile.user_type == 'hod':
                return redirect('hod_dashboard')
            elif profile.user_type == 'finance':
                return redirect('finance_dashboard')            
            else:
                return redirect('dashboard')
        else:
            return render(request, 'dashboard/login.html', {'error': 'Invalid username or password'})
    return render(request, 'dashboard/login.html')

# Logout view
def logout_view(request):
    logout(request)
    return redirect('login')

# Dashboard view
@login_required
def dashboard_view(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.is_coordinator:
            return redirect('coordinator_dashboard')
        else:
            return redirect('employee_dashboard')
    return render(request, 'dashboard/view_kaizen.html')  # Changed from view_kaizen_sheets.html

# views.py
@login_required
def employee_dashboard(request):
    if not request.user.profile.is_employee:
        return HttpResponseForbidden()

    if request.method == 'POST':
        title = request.POST.get('title')
        instance = KaizenSheet.objects.filter(
            title=title, 
            employee=request.user,
            is_temporary=False
        ).first()
        
        if instance and instance.approval_status != 'pending':
            messages.error(request, 'Cannot edit an approved kaizen sheet')
            return redirect('employee_dashboard')
        
        form = KaizenSheetForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            kaizen = form.save(commit=False)
            kaizen.employee = request.user
            kaizen.is_temporary = False
            kaizen.save()
            
            horizontal_deployment = request.POST.get('id_horizontal_deployment') == 'on'
            if horizontal_deployment:
                selected_departments = request.POST.getlist('horizontal_departments')
                # Clear existing deployments
                HorizontalDeployment.objects.filter(kaizen_sheet=kaizen).delete()
                
                # Create new deployments
                for dept in selected_departments:
                    print(f"Creating deployment for department: {dept}")
                    HorizontalDeployment.objects.create(
                        kaizen_sheet=kaizen,
                        department=dept
                    )
            # Handle file uploads
            for field in ['before_improvement_image', 'after_improvement_image', 
                         'standardization_file', 'cost_calculation']:
                if field in request.FILES:
                    setattr(kaizen, field, request.FILES[field])
            
            # Handle impact fields
            impacts = ['safety', 'quality', 'productivity', 'delivery', 
                      'cost', 'morale', 'environment']
            for impact in impacts:
                impact_checked = request.POST.get(f'impacts_{impact}') == 'true'
                setattr(kaizen, f'impacts_{impact}', impact_checked)
                
                if impact_checked:
                    for field_type in ['benefits_description', 'uom', 
                                     'before_implementation', 'after_implementation']:
                        field_name = f'{impact}_{field_type}'
                        if field_name in request.POST:
                            setattr(kaizen, field_name, request.POST.get(field_name))
            
            
            kaizen.save()
            messages.success(request, 'Kaizen sheet saved successfully')
            return redirect('employee_dashboard')
    else:
        form = KaizenSheetForm()

    # Get user's kaizen sheets
    kaizen_sheets = KaizenSheet.objects.filter(
        employee=request.user,
        is_temporary=False
    ).order_by('-created_at')

    # Get available departments for horizontal deployment
    departments = Profile.objects.filter(
        user_type='hod'
    ).exclude(
        department=request.user.profile.department
    ).values_list('department', flat=True).distinct()

    context = {
        'form': form,
        'kaizen_sheets': kaizen_sheets,
        'departments': departments,
        'kaizen_list': kaizen_sheets,
        'pending_sheets': kaizen_sheets.filter(approval_status='pending'),
        'hod_approved_sheets': kaizen_sheets.filter(approval_status='hod_approved'),
        'finance_pending_sheets': kaizen_sheets.filter(approval_status='finance_pending'),
        'finance_approved_sheets': kaizen_sheets.filter(approval_status='finance_approved'),
        'coordinator_approved_sheets': kaizen_sheets.filter(approval_status='coordinator_approved'),
        'impacts': ['safety', 'quality', 'productivity', 'delivery', 
                   'cost', 'morale', 'environment']
    }
    
    return render(request, 'dashboard/employee_dashboard.html', context)

# views.py
# views.py
@login_required
def approve_kaizen(request, kaizen_id):
    if not request.user.profile.is_hod:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
        
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    try:
        sheet.hod_approved = True
        sheet.hod_approved_by = request.user
        sheet.hod_approved_at = timezone.now()
        
        cost_diff = sheet.get_cost_difference()
        
        # Update status based on cost difference
        if cost_diff <= 45000:
            sheet.approval_status = 'completed'
        elif cost_diff <= 100000:
            sheet.approval_status = 'coordinator_pending'
        else:
            sheet.approval_status = 'finance_pending'
            
        sheet.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def view_sheet(request, sheet_id):
    if not request.user.profile.is_hod:
        return HttpResponseForbidden()
        
    sheet = get_object_or_404(KaizenSheet, 
        id=sheet_id,
        horizontal_departments=request.user.profile
    )
    
    return render(request, 'dashboard/view_kaizen.html', {'sheet': sheet})

@login_required
def fetch_kaizen_sheet(request, kaizen_id):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        kaizen = get_object_or_404(KaizenSheet, id=kaizen_id)
        
        # Get deployed departments
        deployed_depts = HorizontalDeployment.objects.filter(
            kaizen_sheet=kaizen
        ).values_list('department', flat=True)

        # Impact fields data
        impact_data = {}
        impacts = ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']
        
        for impact in impacts:
            # Add impact checkbox state
            impact_data[f'impacts_{impact}'] = getattr(kaizen, f'impacts_{impact}', False)
            
            # Add impact field values
            for field_type in ['benefits_description', 'uom', 'before_implementation', 'after_implementation']:
                field_name = f'{impact}_{field_type}'
                impact_data[field_name] = getattr(kaizen, field_name, '')
            
            cost_difference = None
            try:
                if kaizen.cost_before_implementation and kaizen.cost_after_implementation:
                    before_value = float(kaizen.cost_before_implementation)
                    after_value = float(kaizen.cost_after_implementation)
                    cost_difference = after_value - before_value
            except ValueError:
                cost_difference = None  # Handle the case where conversion fails
            # Combine all data
        data = {
            # Basic Information
            'title': kaizen.title,
            'area_implemented': kaizen.area_implemented,
            'start_date': kaizen.start_date.strftime('%Y-%m-%d') if kaizen.start_date else '',
            'end_date': kaizen.end_date.strftime('%Y-%m-%d') if kaizen.end_date else '',
            'team_member2_id': kaizen.team_member2_id,
            'team_member2': kaizen.team_member2,

            # Project Details
            'problem': kaizen.problem,
            'idea_solved': kaizen.idea_solved,
            'standardization': kaizen.standardization,
            'benefits': kaizen.benefits,
            'deployment': kaizen.deployment,

            # Horizontal Deployment
            'horizontal_deployment': bool(deployed_depts),
            'horizontal_departments': list(deployed_depts),

            # Improvement Documentation
            'before_improvement_text': kaizen.before_improvement_text,
            'after_improvement_text': kaizen.after_improvement_text,

            # File URLs
            'before_improvement_image': kaizen.before_improvement_image.url if kaizen.before_improvement_image else '',
            'after_improvement_image': kaizen.after_improvement_image.url if kaizen.after_improvement_image else '',
            'standardization_file': kaizen.standardization_file.url if kaizen.standardization_file else '',
            'cost_calculation': kaizen.cost_calculation.url if kaizen.cost_calculation else '',

            'cost_benefits_description': impact_data.get('cost_benefits_description', ''),
            'cost_uom': impact_data.get('cost_uom', ''),
            'cost_before_implementation': impact_data.get('cost_before_implementation', ''),
            'cost_after_implementation': impact_data.get('cost_after_implementation', ''),
            'cost_difference': cost_difference,
            'cost_calculation': kaizen.cost_calculation.url if kaizen.cost_calculation else '',
        }

        # Add impact data to response
        data.update(impact_data)

        # Debug logging
        print("Fetched Kaizen data:", data)
        print("Deployed departments:", deployed_depts)
        print("Impact data:", impact_data)

        return JsonResponse(data)
    
    return HttpResponseBadRequest("Invalid request")
# views.py
@login_required
def create_temp_kaizen(request):
    if request.method == 'POST':
        form = KaizenSheetForm(request.POST, request.FILES)
        if form.is_valid():
            title = form.cleaned_data['title']
            
            # Check for existing non-temporary sheet
            existing_sheet = KaizenSheet.objects.filter(
                title=title,
                employee=request.user,
                is_temporary=False
            ).first()
            
            if existing_sheet:
                # Return existing sheet ID for download
                return JsonResponse({'kaizen_id': existing_sheet.id})
            else:
                if not title:
                    return JsonResponse({
                        'error': 'Please fill required fields before downloading'
                    }, status=400)
                
                # Create temporary sheet for download
                kaizen = form.save(commit=False)
                kaizen.employee = request.user
                kaizen.is_temporary = True
                kaizen.save()
                
                return JsonResponse({'kaizen_id': kaizen.id})
    
    return JsonResponse({'error': 'Invalid form data'}, status=400)

# views.py
def check_impact_data(kaizen, impact):
    return any([
        getattr(kaizen, f'{impact}_benefits_description', ''),
        getattr(kaizen, f'{impact}_uom', ''),
        getattr(kaizen, f'{impact}_before_implementation', ''),
        getattr(kaizen, f'{impact}_after_implementation', '')
    ])

@login_required
def download_kaizen_sheet(request, kaizen_id):
    try:
        kaizen = get_object_or_404(KaizenSheet, id=kaizen_id, employee=request.user)
        
        # Load template
        template_path = 'kaizen_app/templates/excel/kaizen_format.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active

        def write_to_merged_cell(sheet, cell_address, value, is_yes_no=False, font_style=None):
            for merged_range in sheet.merged_cells.ranges:
                if cell_address in merged_range:
                    sheet.unmerge_cells(str(merged_range))
                    cell = sheet[cell_address]
                    cell.value = value
                    if is_yes_no:
                        cell.font = Font(name='Arial', size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif font_style:
                        cell.font = font_style
                    sheet.merge_cells(str(merged_range))
                    return
                    
            # If not merged, just set the value
            cell = sheet[cell_address]
            cell.value = value
            if is_yes_no:
                cell.font = Font(name='Arial', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif font_style:
                cell.font = font_style

        # Basic info
        write_to_merged_cell(ws, "C2", kaizen.title)
        write_to_merged_cell(ws, "T3", kaizen.serial_key)
        write_to_merged_cell(ws, "A5", kaizen.area_implemented)
        write_to_merged_cell(ws, "E5", kaizen.start_date.strftime('%Y-%m-%d'))
        write_to_merged_cell(ws, "H5", kaizen.end_date.strftime('%Y-%m-%d'))
        write_to_merged_cell(ws, "S4", kaizen.employee.username)

        # Problem section
        write_to_merged_cell(ws, "A7", kaizen.problem)

        # Before improvement section
        if kaizen.before_improvement_image:
            try:
                image_path = os.path.join(settings.MEDIA_ROOT, str(kaizen.before_improvement_image))
                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 300
                    img.height = 200
                    ws.add_image(img, 'G7')
            except Exception as e:
                print(f"Error adding before image: {str(e)}")
        write_to_merged_cell(ws, "G16", kaizen.before_improvement_text, 
                           font_style=Font(name='Arial', size=11, bold=False))

        # Ideas section
        write_to_merged_cell(ws, "A21", kaizen.idea_solved)

        # After improvement section
        if kaizen.after_improvement_image:
            try:
                image_path = os.path.join(settings.MEDIA_ROOT, str(kaizen.after_improvement_image))
                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 300
                    img.height = 200
                    ws.add_image(img, 'G21')
            except Exception as e:
                print(f"Error adding after image: {str(e)}")
        write_to_merged_cell(ws, "G32", kaizen.after_improvement_text, 
                           font_style=Font(name='Arial', size=11, bold=False))

        # Impact matrix with correct cell mappings
        impacts = [
            ('safety', 10),     # P10-V10
            ('quality', 12),    # P12-V12
            ('productivity', 14), # P14-V14
            ('delivery', 16),   # P16-V16
            ('cost', 18),       # P18-V18
            ('morale', 20),     # P20-V20
            ('environment', 22) # P22-V22
        ]

        for impact_name, row_num in impacts:
            # Check if impact has data
            is_checked = check_impact_data(kaizen, impact_name)
            print(f"Impact {impact_name}: {is_checked}")  # Debug log
            
            # Write Yes/No with special handling
            write_to_merged_cell(ws, f"P{row_num}", 'Yes' if is_checked else 'No', is_yes_no=True)
            
            # Write impact data fields
            write_to_merged_cell(ws, f"Q{row_num}", getattr(kaizen, f'{impact_name}_benefits_description', ''))
            write_to_merged_cell(ws, f"S{row_num}", getattr(kaizen, f'{impact_name}_uom', '') or '')
            write_to_merged_cell(ws, f"T{row_num}", getattr(kaizen, f'{impact_name}_before_implementation', ''))
            write_to_merged_cell(ws, f"V{row_num}", getattr(kaizen, f'{impact_name}_after_implementation', ''))

        # Standardization and deployment
        write_to_merged_cell(ws, "N25", kaizen.standardization)
        write_to_merged_cell(ws, "N32", kaizen.deployment)

        # Set response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Kaizen_Sheet_{kaizen.serial_key}.xlsx'
        wb.save(response)
        return response

    except Exception as e:
        print(f"Error in download_kaizen_sheet: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def edit_kaizen_sheet(request, sheet_id):
    kaizen_sheet = get_object_or_404(KaizenSheet, id=sheet_id, employee=request.user)
    
    if kaizen_sheet.is_approved:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = KaizenSheetForm(request.POST, request.FILES, instance=kaizen_sheet)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = KaizenSheetForm(instance=kaizen_sheet)

    return render(request, 'dashboard/edit_kaizen_sheet.html', {'form': form, 'sheet_id': sheet_id})

@login_required
def save_kaizen_sheet(request):
    if request.method == 'POST':
        form = KaizenSheetForm(request.POST, request.FILES)
        if form.is_valid():
            kaizen_sheet = form.save(commit=False)
            kaizen_sheet.employee = request.user
            kaizen_sheet.save()
            
            if 'before_improvement_image' in request.FILES:
                img = Image.objects.create(image=request.FILES['before_improvement_image'])
                kaizen_sheet.before_improvement_images.add(img)
            
            if 'after_improvement_image' in request.FILES:
                img = Image.objects.create(image=request.FILES['after_improvement_image'])
                kaizen_sheet.after_improvement_images.add(img)
            
            kaizen_sheet.save()
            return redirect('employee_dashboard')
    else:
        form = KaizenSheetForm()
    return render(request, 'dashboard/employee_dashboard.html', {'form': form})

# views.py
@login_required
def approve_kaizen(request, kaizen_id):
    if not request.user.profile.is_hod:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
        
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    try:
        sheet.hod_approved = True
        sheet.hod_approved_by = request.user
        sheet.hod_approved_at = timezone.now()
        
        cost_diff = sheet.get_cost_difference()
        
        # Update status based on cost difference
        if cost_diff <= 45000:
            sheet.approval_status = 'completed'
        elif cost_diff <= 100000:
            sheet.approval_status = 'coordinator_pending'
        else:
            sheet.approval_status = 'finance_pending'
            
        sheet.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
def get_department_stats(dept_name):
    """Helper function to get department statistics"""
    if not dept_name:
        return None
        
    hod = User.objects.filter(
        profile__department=dept_name,
        profile__user_type='hod'
    ).first()

    employee_count = Profile.objects.filter(
        department=dept_name,
        user_type='employee'
    ).count()

    try:
        coordinator_entry = KaizenCoordinator.objects.get_or_create(
            department=dept_name,
            defaults={'coordinator_name': ''}
        )[0]
        coordinator_name = coordinator_entry.coordinator_name or ""
    except Exception:
        coordinator_name = ""

    completed_sheets = KaizenSheet.objects.filter(
        employee__profile__department=dept_name,
        approval_status='completed',
        hod_approved_at__isnull=False
    ).count()

    monthly_target = employee_count
    achievement_rate = (completed_sheets / monthly_target * 100) if monthly_target > 0 else 0

    return {
        'name': dept_name,
        'hod_name': hod.username if hod else "No HOD Assigned",
        'kaizen_coordinator': coordinator_name,
        'employee_count': employee_count,
        'monthly_target': monthly_target,
        'completed': completed_sheets,
        'achievement_rate': round(achievement_rate, 2),
    }

@login_required
def coordinator_dashboard(request):
    if not request.user.profile.is_coordinator:
        return HttpResponseForbidden()

    # Get latest timestamp and date calculations
    latest_sheet = KaizenSheet.objects.order_by('-created_at').first()
    today = timezone.localtime(latest_sheet.created_at if latest_sheet else timezone.now())
    current_month = today.month
    current_year = today.year

    # Academic year calculations
    start_year = current_year if current_month >= 4 else current_year - 1
    academic_years_data = []
    
    for year in range(start_year - 4, start_year + 1):
        start_date = timezone.make_aware(timezone.datetime(year, 4, 1))
        end_date = timezone.make_aware(timezone.datetime(year + 1, 3, 31))
        
        if end_date > today:
            end_date = today
            
        completed_sheets = KaizenSheet.objects.filter(
            hod_approved_at__range=(start_date, end_date),
            approval_status='completed'
        )
        
        months_diff = ((end_date.year - start_date.year) * 12 + 
                    end_date.month - start_date.month + 1)
                    
        monthly_average = round(completed_sheets.count() / months_diff, 2) if months_diff > 0 else 0
        
        academic_years_data.append({
            'year': f'AY {year}-{year + 1}',
            'average': monthly_average
        })

    # Monthly calculations
    monthly_submissions = [0] * 12
    monthly_completions = [0] * 12
    monthly_averages = [0] * 12
    cumulative_completed = 0
   
    total_employees = User.objects.filter(profile__user_type='employee').count()

    # Calculate monthly statistics
    for month in range(1, 13):
        month_start = timezone.make_aware(timezone.datetime(current_year, month, 1))
        month_end = timezone.make_aware(
            (timezone.datetime(current_year, month + 1, 1) if month < 12 
             else timezone.datetime(current_year + 1, 1, 1)) - timezone.timedelta(days=1)
        )

        submissions = KaizenSheet.objects.filter(
            created_at__range=(month_start, month_end)
        ).count()

        completions = KaizenSheet.objects.filter(
            hod_approved_at__range=(month_start, month_end),
            approval_status='completed'
        ).count()

        monthly_submissions[month - 1] = submissions
        monthly_completions[month - 1] = completions

        if month_start <= today:
            cumulative_completed += completions
            monthly_averages[month - 1] = round(cumulative_completed / month, 2)
        else:
            monthly_averages[month - 1] = None

    # Department statistics
    departments = []
    department_names = Profile.objects.values_list('department', flat=True).distinct().exclude(department__isnull=True)

    for dept_name in department_names:
        dept_data = get_department_stats(dept_name)
        if dept_data:
            departments.append(dept_data)

    # Employee statistics
    employee_stats = []
    for user in User.objects.filter(profile__isnull=False).select_related('profile'):
        user_sheets = KaizenSheet.objects.filter(employee=user)
        stats = {
            'id': user.id,
            'username': user.username,
            'department': user.profile.department,
            'total_submissions': user_sheets.count(),
            'pending': user_sheets.filter(approval_status='pending').count(),
            'approved': user_sheets.count() - user_sheets.filter(approval_status='pending').count()
        }
        employee_stats.append(stats)

    context = {
        'current_year': current_year,
        'year_range': range(current_year, current_year - 5, -1),
        'months': json.dumps(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']),
        'monthly_submissions': json.dumps(monthly_submissions, cls=DjangoJSONEncoder),
        'monthly_completions': json.dumps(monthly_completions, cls=DjangoJSONEncoder),
        'monthly_averages': json.dumps(monthly_averages, cls=DjangoJSONEncoder),
        'academic_years': json.dumps([d['year'] for d in academic_years_data]),
        'academic_year_counts': json.dumps([d['average'] for d in academic_years_data], cls=DjangoJSONEncoder),
        'current_academic_year': f"{start_year}-{start_year + 1}",
        'current_month': current_month,
        'total_employees': total_employees,
        'departments': departments,
        'total_kaizens': KaizenSheet.objects.count(),
        'completed_count': KaizenSheet.objects.filter(approval_status='completed').count(),
        'hod_pending_count': KaizenSheet.objects.filter(approval_status='pending').count(),
        'coordinator_pending_count': KaizenSheet.objects.filter(approval_status='hod_approved').count(),
        'finance_pending_count': KaizenSheet.objects.filter(approval_status='finance_pending').count(),
        'pending_approvals': KaizenSheet.objects.exclude(approval_status='completed').count(),
        'coordinator_approved': KaizenSheet.objects.filter(
            models.Q(approval_status='completed') |
            models.Q(approval_status='coordinator_approved')
        ).count(),
        'employee_stats': employee_stats,
        'department_data': json.dumps(departments, cls=DjangoJSONEncoder)
    }

    return render(request, 'dashboard/coordinator_dashboard.html', context)

@login_required
@require_POST
def save_kaizen_coordinators(request):
    if not request.user.profile.is_coordinator:
        return JsonResponse({'success': False, 'error': 'Access denied.'})

    try:
        data = json.loads(request.body)
        for department, coordinator_name in data.items():
            KaizenCoordinator.objects.update_or_create(
                department=department,
                defaults={'coordinator_name': coordinator_name}
            )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_yearly_data(request, year):
    if not request.user.profile.is_coordinator:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        year = int(year)
        today = timezone.localtime(timezone.now())
        total_employees = User.objects.filter(profile__user_type='employee').count()

        # Initialize arrays
        monthly_submissions = [0] * 12
        monthly_completions = [0] * 12

        # Process each month
        for month in range(1, 13):
            # Create timezone-aware month boundaries
            month_start = timezone.make_aware(datetime(year, month, 1))
            month_end = timezone.make_aware(
                (datetime(year, month + 1, 1) if month < 12 
                 else datetime(year + 1, 1, 1)) - timedelta(days=1)
            )

            # Skip future months
            if month_start > today:
                continue

            # Get submissions for the month
            submissions = KaizenSheet.objects.filter(
                created_at__range=(month_start, month_end)
            ).count()
            monthly_submissions[month - 1] = submissions

            # Get completions for the month
            completions = KaizenSheet.objects.filter(
                approval_status='completed',
                hod_approved_at__range=(month_start, month_end)
            ).count()
            monthly_completions[month - 1] = completions

        return JsonResponse({
            'months': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'submissions': monthly_submissions,
            'completions': monthly_completions
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_department_data(request, year, month):
    if not request.user.profile.is_coordinator:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        # Convert month and year to date range
        start_date = timezone.make_aware(datetime(year, month, 1))
        if month == 12:
            end_date = timezone.make_aware(datetime(year + 1, 1, 1)) - timezone.timedelta(days=1)
        else:
            end_date = timezone.make_aware(datetime(year, month + 1, 1)) - timezone.timedelta(days=1)

        departments = []
        department_names = Profile.objects.values_list('department', flat=True).distinct()

        for dept_name in department_names:
            # Get HOD
            hod = User.objects.filter(
                profile__department=dept_name,
                profile__user_type='hod'
            ).first()

            # Employee count
            employee_count = Profile.objects.filter(
                department=dept_name,
                user_type='employee'
            ).count()

            # Get coordinator
            kaizen_coordinator_entry = KaizenCoordinator.objects.filter(department=dept_name).first()
            coordinator_name = kaizen_coordinator_entry.coordinator_name if kaizen_coordinator_entry else ""

            # Get completed sheets for the selected month
            completed_sheets = KaizenSheet.objects.filter(
                employee__profile__department=dept_name,
                approval_status='completed',
                hod_approved_at__range=(start_date, end_date)
            ).count()

            # Calculate achievement rate
            monthly_target = employee_count
            achievement_rate = (completed_sheets / monthly_target * 100) if monthly_target > 0 else 0

            dept_data = {
                'name': dept_name,
                'hod_name': hod.username if hod else "No HOD Assigned",
                'kaizen_coordinator': coordinator_name,
                'employee_count': employee_count,
                'monthly_target': monthly_target,
                'completed': completed_sheets,
                'achievement_rate': round(achievement_rate, 2)
            }
            departments.append(dept_data)

        return JsonResponse({
            'success': True,
            'departments': departments
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_excel_template(request, template_name):
    try:
        file_path = os.path.join(settings.BASE_DIR, 'kaizen_app', 'templates', 'excel', template_name)
        if os.path.exists(file_path):
            with open(file_path, 'rb') as template:
                response = HttpResponse(
                    template.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{template_name}"'
                return response
        return HttpResponse('Template not found', status=404)
    except Exception as e:
        return HttpResponse(str(e), status=500)

# views.py
@login_required
def cip_register_view(request):
    print("CIP Register view accessed") # Debug print
    if not request.user.profile.is_coordinator:
        return HttpResponseForbidden("Access denied. Coordinator privileges required.")
        
    try:
        kaizen_sheets = KaizenSheet.objects.select_related('employee').all().order_by('-created_at')
        
        processed_sheets = []
        impact_types = ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']
        
        for sheet in kaizen_sheets:

            department = sheet.employee.profile.department
            hod = User.objects.filter(
                profile__user_type='hod',
                profile__department=department
            ).first()


            sheet_data = {
                # Basic fields
                'id': sheet.id,
                'serial_key': sheet.serial_key,
                'title': sheet.title,
                'employee': sheet.employee,
                'created_at': sheet.created_at,
                'approval_status': sheet.approval_status,
                'area_implemented': sheet.area_implemented,
                'start_date': sheet.start_date,
                'end_date': sheet.end_date,
                
                # New team member fields
                'project_leader': hod.username if hod else "No HOD Assigned",
                'team_member1_id': sheet.team_member1_id,
                'team_member1': sheet.team_member1,
                'team_member2_id': sheet.team_member2_id,
                'team_member2': sheet.team_member2,
                
                # New savings fields
                'savings_start_month': sheet.savings_start_month,
                'estimated_savings': sheet.estimated_savings,
                'realized_savings': sheet.realized_savings,

                # File fields with proper URL handling
                'standardization_file': sheet.standardization_file,
                'cost_calculation': sheet.cost_calculation,
                'before_improvement_image': sheet.before_improvement_image,
                'after_improvement_image': sheet.after_improvement_image,
                
                # Handwritten sheet fields
                'is_handwritten': sheet.is_handwritten,
                'handwritten_sheet': sheet.handwritten_sheet,
                
                # Processed URLs
                'standardization_file_url': sheet.standardization_file.url if sheet.standardization_file else None,
                'cost_calculation_url': sheet.cost_calculation.url if sheet.cost_calculation else None,
                'before_improvement_image_url': sheet.before_improvement_image.url if sheet.before_improvement_image else None,
                'after_improvement_image_url': sheet.after_improvement_image.url if sheet.after_improvement_image else None,
                'handwritten_sheet_url': sheet.handwritten_sheet.url if sheet.handwritten_sheet else None,
                
                # Container for impact data
                'impact_data': {}
            }
            
            # Process impact data for each type
            for impact in impact_types:
                sheet_data['impact_data'][impact] = {
                    'benefits_description': getattr(sheet, f'{impact}_benefits_description', ''),
                    'uom': getattr(sheet, f'{impact}_uom', ''),
                    'before_implementation': getattr(sheet, f'{impact}_before_implementation', ''),
                    'after_implementation': getattr(sheet, f'{impact}_after_implementation', ''),
                    'impacts': getattr(sheet, f'impacts_{impact}', False)
                }
            
            processed_sheets.append(sheet_data)

        context = {
            'kaizen_sheets': processed_sheets,
            'impacts': impact_types,
            'user': request.user,
            'MEDIA_URL': settings.MEDIA_URL
        }
        
        return render(request, 'dashboard/cip_register.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading CIP register: {str(e)}')
        return redirect('coordinator_dashboard')

# views.py - Update finance approval view
@login_required
def finance_approve_kaizen(request, kaizen_id):
    if not request.user.profile.user_type == 'finance':
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    action = request.POST.get('action')
    
    try:
        if action == 'approve':
            sheet.finance_approved = True
            sheet.finance_approved_by = request.user
            sheet.finance_approved_at = timezone.now()
            sheet.approval_status = 'coordinator_pending'
        elif action == 'reject':
            sheet.finance_approved = False
            sheet.approval_status = 'finance_rejected'
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        sheet.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
       
@login_required
def coordinator_approve_kaizen(request, kaizen_id):
    if not request.user.profile.is_coordinator:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
        
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    try:
        if sheet.approval_status != 'coordinator_pending':
            return JsonResponse({'success': False, 'error': 'Invalid sheet status'})
            
        sheet.coordinator_approved = True
        sheet.coordinator_approved_by = request.user
        sheet.coordinator_approved_at = timezone.now()
        sheet.approval_status = 'completed'
        sheet.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def view_kaizen(request, kaizen_id):
    if not request.user.profile.is_hod:
        return HttpResponseForbidden()
    
    department = request.user.profile.department
    
    # First try to get sheet from HOD's department
    try:
        sheet = KaizenSheet.objects.select_related('employee__profile').get(
            id=kaizen_id,
            employee__profile__department=department
        )

            # Create impact data dictionary from direct fields
        is_horizontal = False
    except KaizenSheet.DoesNotExist:
        # If not found, check horizontal deployments
        sheet = get_object_or_404(
            KaizenSheet.objects.select_related('employee__profile'),
            id=kaizen_id,
            deployments__department=department  # Using related_name from HorizontalDeployment
        )
        is_horizontal = True

    impact_types = ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']
    impact_data = {}
    
    for impact in impact_types:
        impact_data[impact] = {
            'benefits_description': getattr(sheet, f'{impact}_benefits_description', ''),
            'uom': getattr(sheet, f'{impact}_uom', ''),
            'before_implementation': getattr(sheet, f'{impact}_before_implementation', ''),
            'after_implementation': getattr(sheet, f'{impact}_after_implementation', '')
        }

    context = {
        'sheet': sheet,
        'is_horizontal': is_horizontal,
        'source_department': sheet.employee.profile.department,
        'impacts': impact_types,
        'impact_data': impact_data,
        'is_hod': request.user.profile.is_hod,
        'is_coordinator': request.user.profile.is_coordinator        
    }
    
    return render(request, 'dashboard/view_kaizen.html', context)

@login_required
def get_cost_details(request, kaizen_id):
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    
    data = {
        'benefits_description': sheet.cost_benefits_description,
        'uom': sheet.cost_uom,
        'before_implementation': sheet.cost_before_implementation,
        'after_implementation': sheet.cost_after_implementation,
        'cost_difference': sheet.get_cost_difference(),
        'cost_calculation': sheet.cost_calculation.url if sheet.cost_calculation else None
    }
    
    return JsonResponse(data)

@login_required
def update_kaizen(request, kaizen_id):
    if not request.user.profile.is_coordinator:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            kaizen = KaizenSheet.objects.get(id=kaizen_id)
            
            # List of fields that can be updated
            allowed_fields = [
                'title', 'area_implemented', 'start_date', 'end_date',
                'project_leader', 'team_member1_id', 'team_member1',
                'team_member2_id', 'team_member2', 'savings_start_month',
                'estimated_savings', 'realized_savings'
            ]
            
            # Update fields
            for field, value in data.items():
                if field in allowed_fields and hasattr(kaizen, field):
                    # Convert string numbers to decimal for savings fields
                    if field in ['estimated_savings', 'realized_savings']:
                        try:
                            value = float(value) if value else 0
                        except ValueError:
                            continue
                    setattr(kaizen, field, value)
            
            kaizen.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'})

# views.py
@login_required
def hod_dashboard(request):
    if request.user.profile.user_type != 'hod':
        return HttpResponseForbidden()
    
    department = request.user.profile.department
    current_year = timezone.now().year
    year_range = range(current_year, current_year - 5, -1)

    # Department stats for pie chart
    dept_kaizens = KaizenSheet.objects.filter(employee__profile__department=department)
    dept_completed = dept_kaizens.filter(approval_status='completed').count()
    dept_hod_pending = dept_kaizens.filter(approval_status='pending').count()
    dept_coordinator_pending = dept_kaizens.filter(approval_status='coordinator_pending').count()
    dept_finance_pending = dept_kaizens.filter(approval_status='finance_pending').count()

    # Get department employees
    department_employees = User.objects.filter(profile__department=department)

    horizontal_sheets = KaizenSheet.objects.filter(
        deployments__department=department
    ).exclude(
        employee__profile__department=department
    ).distinct().order_by('-created_at')    

    context = {
        'kaizen_sheets': dept_kaizens.order_by('-created_at'),
        'horizontal_sheets': horizontal_sheets,
        'dept_completed': dept_completed,
        'dept_hod_pending': dept_hod_pending,
        'dept_coordinator_pending': dept_coordinator_pending,
        'dept_finance_pending': dept_finance_pending,
        'department_employees': department_employees,
        'year_range': year_range
    }
    
    return render(request, 'dashboard/hod_dashboard.html', context)
# Add new view for employee submissions data
@login_required
def get_employee_submissions(request, employee_id):
    # Update permission check to allow coordinators
    if not (request.user.profile.is_coordinator or request.user.profile.is_hod):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        employee = get_object_or_404(User, id=employee_id)
        selected_year = request.GET.get('year', timezone.now().year)
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_data = [0] * 12
        
        # Get submissions for selected year
        submissions = KaizenSheet.objects.filter(
            employee=employee,
            created_at__year=selected_year
        )
        
        # Count submissions by month
        for submission in submissions:
            month_idx = submission.created_at.month - 1
            monthly_data[month_idx] += 1
        
        print(f"Employee {employee.username} data for {selected_year}:", monthly_data)  # Debug log
        
        return JsonResponse({
            'submissions': monthly_data,
            'months': months
        })
        
    except Exception as e:
        print(f"Error in get_employee_submissions: {e}")  # Debug log
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def upload_handwritten_sheet(request):
    if request.method == 'POST' and request.POST.get('form_type') == 'handwritten':
        try:
            title = request.POST.get('handwritten_title')
            handwritten_sheet = request.FILES.get('handwritten_sheet')
            
            if not title or not handwritten_sheet:
                error_msg = []
                if not title:
                    error_msg.append("Title is required")
                if not handwritten_sheet:
                    error_msg.append("Handwritten sheet is required")
                    
                return JsonResponse({
                    'success': False,
                    'error': ' and '.join(error_msg)
                })
                
            kaizen = KaizenSheet(
                title=title,
                employee=request.user,
                is_handwritten=True,
                handwritten_sheet=handwritten_sheet,
                area_implemented="Handwritten Sheet",
                start_date=timezone.now().date(),
                end_date=timezone.now().date(),
                problem="See handwritten sheet",
                idea_solved="See handwritten sheet",
                standardization="See handwritten sheet",
                benefits="See handwritten sheet",
                deployment="See handwritten sheet",
                approval_status='pending'  # Always starts at pending for HOD approval
            )
            kaizen.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Handwritten sheet uploaded successfully',
                'serial_key': kaizen.serial_key
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method or form type'})
@login_required
def finance_dashboard(request):
    if not request.user.profile.user_type == 'finance':
        return HttpResponseForbidden()

    # Only get kaizens that need finance approval (>1 lakh)
    high_value_kaizens = KaizenSheet.objects.filter(
        models.Q(cost_before_implementation__gt=100000) |
        models.Q(cost_after_implementation__gt=100000)
    ).order_by('-created_at')

    context = {
        'kaizens': high_value_kaizens,
        'total_count': high_value_kaizens.count(),
        'pending_count': high_value_kaizens.filter(approval_status='finance_pending').count(),
        'approved_count': high_value_kaizens.filter(
            models.Q(approval_status='completed') |
            models.Q(approval_status='coordinator_pending')
        ).count()
    }

    return render(request, 'dashboard/finance_dashboard.html', context)